import tkinter as tk
from tkinter import *
from tkinter import messagebox
from tkinter import filedialog as fd
from tkinter import ttk as ttk
from PIL import ImageTk,Image
from PIL import Image
import pandas as pd
import openpyxl
import random



# Page Close Confirmations (Messagebox):

global playG,point,instruc

def homelogout():
    messagebox.showinfo('Thank You','See you later')
    home.destroy()
        
def homeclose():
    if messagebox.askokcancel('Quit','Do you want to quit?'):
        home.destroy()
        quit()
        
def playGclose():
    playG.destroy()
 
def playMclose():
    playM.destroy()

def pointclose():
    point.destroy()

def instrucclose():
    instruc.destroy()




# # icon window

icon = tk.Tk()
icon.title('Monty Hall')
icon.iconbitmap("icon ICO.ico")
image = Image.open("icon.png")
tk_image = ImageTk.PhotoImage(image)
image_label = tk.Label(icon, image=tk_image)
image_label.pack()
icon.update()
screen_width = icon.winfo_screenwidth()
screen_height = icon.winfo_screenheight()
window_width = 366  
window_height = 612
x = int((screen_width - window_width) / 2)
y = int((screen_height - window_height) / 2)
icon.geometry("+{}+{}".format(x, y))
icon.after(2000, icon.destroy)
icon.mainloop()





#Excel add
ex_filename = "Monty Hall Points.xlsx"






#Play:

def swap_door():
    global choosen_doors,car_selected_door,dum_door,Choose_door_button1,Choose_door_button2,ch_door2
    if choosen_doors==car_selected_door:
        choosen_doors=dum_door
    else:
        choosen_doors=car_selected_door
        if choosen_doors==car_selected_door:
            wb = openpyxl.load_workbook(ex_filename)
            sheet = wb.active
            row_num = 2
            col_num = 1
            old_value = sheet.cell(row=row_num, column=col_num).value
            new_value = old_value +1
            sheet.cell(row=row_num, column=col_num).value = new_value
            col_num = 2
            old_value = sheet.cell(row=row_num, column=col_num).value
            new_value = old_value +1
            sheet.cell(row=row_num, column=col_num).value = new_value
            col_num = 4
            old_value = sheet.cell(row=row_num, column=col_num).value
            new_value = old_value +1
            sheet.cell(row=row_num, column=col_num).value = new_value
            col_num = 5
            old_value = sheet.cell(row=row_num, column=col_num).value
            new_value = old_value +1
            sheet.cell(row=row_num, column=col_num).value = new_value
            wb.save(ex_filename)
            messagebox.showinfo('You WIN', f'''Win Door: {car_selected_door} 
Swaped Door: {choosen_doors}''')
        else:
            wb = openpyxl.load_workbook(ex_filename)
            sheet = wb.active
            row_num = 2
            col_num = 1
            old_value = sheet.cell(row=row_num, column=col_num).value
            new_value = old_value +1
            sheet.cell(row=row_num, column=col_num).value = new_value
            col_num = 3
            old_value = sheet.cell(row=row_num, column=col_num).value
            new_value = old_value +1
            sheet.cell(row=row_num, column=col_num).value = new_value
            col_num = 4
            old_value = sheet.cell(row=row_num, column=col_num).value
            new_value = old_value +1
            sheet.cell(row=row_num, column=col_num).value = new_value
            wb.save(ex_filename)
            messagebox.showinfo('You LOSE', f'''Win Door: {car_selected_door} 
Swaped Door: {choosen_doors}''')
    Choose_door_button1.destroy()
    Choose_door_button2.destroy()
    ch_door2.destroy()
      
      
      
      
      
        
        
    
    
def stay_door():    
    global choosen_doors,car_selected_door,dum_door,Choose_door_button1,Choose_door_button2,ch_door2
    choosen_doors=choosen_doors
    if choosen_doors==car_selected_door:
        wb = openpyxl.load_workbook(ex_filename)
        sheet = wb.active
        row_num = 2
        col_num = 1
        old_value = sheet.cell(row=row_num, column=col_num).value
        new_value = old_value +1
        sheet.cell(row=row_num, column=col_num).value = new_value
        col_num = 2
        old_value = sheet.cell(row=row_num, column=col_num).value
        new_value = old_value +1
        sheet.cell(row=row_num, column=col_num).value = new_value
        col_num = 6
        old_value = sheet.cell(row=row_num, column=col_num).value
        new_value = old_value +1
        sheet.cell(row=row_num, column=col_num).value = new_value
        wb.save(ex_filename)
        messagebox.showinfo('You WIN', f'''Win Door: {car_selected_door} 
Stayed Door: {choosen_doors}''')
    else:
        wb = openpyxl.load_workbook(ex_filename)
        sheet = wb.active
        row_num = 2
        col_num = 3
        old_value = sheet.cell(row=row_num, column=col_num).value
        new_value = old_value +1
        sheet.cell(row=row_num, column=col_num).value = new_value
        wb.save(ex_filename)
        messagebox.showinfo('You LOSE', f'''Win Door: {car_selected_door} 
Stayed Door: {choosen_doors}''')
    Choose_door_button1.destroy()
    Choose_door_button2.destroy()
    ch_door2.destroy()
    
    
    
    
    
    
    
def startgame():
    global playG,entry1,no_of_doors
    no_of_doors=int(entry1.get())
    if int(no_of_doors) > 30 or int(no_of_doors) <= 2:
        messagebox.showerror('Invalid','Specify Valid Number Of Doors')
    else:
        playG.destroy()
        montyGame()
        
        
        
        
        
        
       
       
        
        
def montyGame():
    
    global playM,entry2,choosen_doors,no_of_doors,ch_door,Choose_door_button,car_selected_door,dum_door

    playM=tk.Toplevel()
    playM.iconbitmap("icon ICO.ico")
    playM.title('Monty Hall')
    playM.state('zoomed')
    playM.protocol('WM_DELETE_WINDOW',playMclose)
    playMpic=ImageTk.PhotoImage(Image.open("Game image.png"))
    playMpicpanel=Label(playM,image=playMpic)
    playMpicpanel.pack(side='top',fill='both',expand='yes')

    def Result():
        global playM,entry2,choosen_doors,Choose_door_button,ch_door,dum_door,Choose_door_button1,Choose_door_button2,ch_door2
        choosen_doors=int(entry2.get())
            
        listdoor=[]
        for i in range(1,no_of_doors+1):
            listdoor.append(i)
        if choosen_doors==car_selected_door:
            listdoor.remove(choosen_doors)
            dum_door = random.randint(1, no_of_doors)
            while dum_door == choosen_doors:
                dum_door = random.randint(1, no_of_doors)
            listdoor.remove(dum_door)
        else:
            listdoor.remove(car_selected_door)
            listdoor.remove(choosen_doors)
            
        #messagebox.showinfo('Door Selected', f'list: {listdoor}')
        
        ch_door.destroy()
        entry2.destroy()
        Choose_door_button.destroy()
        if choosen_doors==car_selected_door:
            ch_door2=tk.Label(playM,text=f"Current Door {choosen_doors}, Optional Door {dum_door}",font=('Comic Sans MS',12,'bold'),height=2,width=28,bg='#3c1c0f',fg='white')
            ch_door2.place(relx=0.15,rely=0.88,anchor=CENTER)
        else:
            ch_door2=tk.Label(playM,text=f"Current Door {choosen_doors}, Optional Door {car_selected_door}",font=('Comic Sans MS',12,'bold'),height=2,width=28,bg='#3c1c0f',fg='white')
            ch_door2.place(relx=0.15,rely=0.88,anchor=CENTER)
        Choose_door_button1=tk.Button(playM,text='Swap',font=('Comic Sans MS',14,'bold'),command=swap_door,height=1,width=9,bg='#3c1c0f',
        fg='white',activebackground='black',activeforeground='white')
        Choose_door_button1.place(relx=0.1,rely=0.95,anchor=CENTER)
        Choose_door_button2=tk.Button(playM,text='Stay',font=('Comic Sans MS',14,'bold'),command=stay_door,height=1,width=9,bg='#3c1c0f',
        fg='white',activebackground='black',activeforeground='white')
        Choose_door_button2.place(relx=0.2,rely=0.95,anchor=CENTER)
        
    
    car_selected_door = random.randint(1, no_of_doors)
    
    ch_door=tk.Label(playM,text="Choose Door",font=('Comic Sans MS',14,'bold'),height=2,width=14,bg='#3c1c0f',fg='white')
    ch_door.place(relx=0.1,rely=0.85,anchor=CENTER)
    def on_validate(*args):
        try:
            
            int_value = int(entry_var.get())
            if 0 < int_value <= no_of_doors:
                entry_var.set(int_value)
            else:
                entry_var.set(no_of_doors)
        except ValueError:
            entry_var.set('')
            
    entry_var = tk.StringVar(value='3')
    entry_var.trace('w', on_validate)
    entry2 = tk.Entry(playM, validate="key", validatecommand=(playM.register(on_validate), '%d', '%P'), textvariable=entry_var)
    entry2.place(relx=0.1,rely=0.9,anchor=CENTER)
    entry2.config(width=28,borderwidth=2,relief='sunken',bg="white",fg="#3c1c0f")
    Choose_door_button=tk.Button(playM,text='Submit',font=('Comic Sans MS',14,'bold'),command=Result,height=1,width=9,bg='#3c1c0f',
        fg='white',activebackground='black',activeforeground='white')
    Choose_door_button.place(relx=0.1,rely=0.95,anchor=CENTER)
    
    
    playM.mainloop()
    


    
    
    
    
    
    
    
def playGame():
    
    global playG,entry1
    
    playG=tk.Toplevel()
    playG.iconbitmap("icon ICO.ico")
    playG.title('Monty Hall')
    playG.configure(bg="black")
    playG.state('zoomed')
    playG.protocol('WM_DELETE_WINDOW',playGclose)

    playGpic=ImageTk.PhotoImage(Image.open("Front image.png"))
    playGpicpanel=Label(playG,image=playGpic)
    playGpicpanel.pack(side='top',fill='both',expand='yes')
    
    def on_validate(*args):
        try:
            int_value = int(entry_var.get())
            if 3 <= int_value <= 30:
                entry_var.set(int_value)
            else:
                entry_var.set('30')
        except ValueError:
            entry_var.set('')

    Label(playG,text="""Number Of Doors
(Max 30)""",font=('Comic Sans MS',14,'bold'),height=2,width=14,bg='#3c1c0f',fg='white').place(relx=0.5,rely=0.45,anchor=CENTER)
    entry_var = tk.StringVar(value='3')
    entry_var.trace('w', on_validate)
    entry1 = tk.Entry(playG, validate="key", validatecommand=(playG.register(on_validate), '%d', '%P'), textvariable=entry_var)
    entry1.place(relx=0.5,rely=0.5,anchor=CENTER)
    entry1.config(width=28,borderwidth=2,relief='sunken',bg="white",fg="#3c1c0f")
    Button(playG,text='Start',font=('Comic Sans MS',14,'bold'),command=startgame,height=1,width=9,bg='#3c1c0f',
       fg='white',activebackground='black',activeforeground='white').place(relx=0.5,rely=0.55,anchor=CENTER)

    playG.mainloop()
    
    
    
    
    
    
    
    

# Points Page:

def pointspage():
    global point
    point=tk.Toplevel()
    
    point.title('Monty Hall - Points')
    point.iconbitmap("icon ICO.ico")
    point.state('zoomed')
    point.protocol('WM_DELETE_WINDOW',pointclose)

    pointpic=ImageTk.PhotoImage(Image.open("Front image.png"))
    pointpanel=Label(point,image=pointpic)
    pointpanel.pack(side='top',fill='both',expand='yes')

    df = pd.read_excel("Monty Hall Points.xlsx")
    point_values=[]
    for index, row in df.iterrows():
        point_values.extend(list(tuple(row)))

    Label(point,text='Games',font=('Comic Sans MS',14,'bold'),height=1,width=14,bg='#3c1c0f',
       fg='white').place(relx=0.155,rely=0.45,anchor=CENTER)
    Label(point,text='Wins',font=('Comic Sans MS',14,'bold'),height=1,width=14,bg='#3c1c0f',
       fg='white').place(relx=0.383,rely=0.45,anchor=CENTER)
    Label(point,text='Loses',font=('Comic Sans MS',14,'bold'),height=1,width=14,bg='#3c1c0f',
       fg='white').place(relx=0.619,rely=0.45,anchor=CENTER)
    
    Label(point,text=point_values[0],font=('Comic Sans MS',14,'bold'),height=1,width=14,bg='white',
       fg='#3c1c0f').place(relx=0.155,rely=0.5,anchor=CENTER)
    Label(point,text=point_values[1],font=('Comic Sans MS',14,'bold'),height=1,width=14,bg='white',
       fg='#3c1c0f').place(relx=0.383,rely=0.5,anchor=CENTER)
    Label(point,text=point_values[2],font=('Comic Sans MS',14,'bold'),border=1,height=1,width=14,bg='white',
       fg='#3c1c0f').place(relx=0.619,rely=0.5,anchor=CENTER)
    
    
    Label(point,text='Total Swaps',font=('Comic Sans MS',14,'bold'),height=1,width=14,bg='#3c1c0f',
       fg='white').place(relx=0.155,rely=0.6,anchor=CENTER)
    Label(point,text='Swaped to win',font=('Comic Sans MS',14,'bold'),height=1,width=14,bg='#3c1c0f',
       fg='white').place(relx=0.383,rely=0.6,anchor=CENTER)
    Label(point,text='Stayed to Win',font=('Comic Sans MS',14,'bold'),height=1,width=14,bg='#3c1c0f',
       fg='white').place(relx=0.619,rely=0.6,anchor=CENTER)
    
    Label(point,text=point_values[3],font=('Comic Sans MS',14,'bold'),height=1,width=14,bg='white',
       fg='#3c1c0f').place(relx=0.155,rely=0.65,anchor=CENTER)
    Label(point,text=point_values[4],font=('Comic Sans MS',14,'bold'),height=1,width=14,bg='white',
       fg='#3c1c0f').place(relx=0.383,rely=0.65,anchor=CENTER)
    Label(point,text=point_values[5],font=('Comic Sans MS',14,'bold'),border=1,height=1,width=14,bg='white',
       fg='#3c1c0f').place(relx=0.619,rely=0.65,anchor=CENTER)


    def reset():
        for i in range(1,7):
            wb = openpyxl.load_workbook(ex_filename)
            sheet = wb.active
            row_num = 2
            col_num = i
            new_value = 0
            sheet.cell(row=row_num, column=col_num).value = new_value
            wb.save(ex_filename)
            point.destroy()   

    Button(point,text='Reset points',font=('Comic Sans MS',18),command=reset,height=1,width=10,bg='red',
       fg='white',activebackground='black',activeforeground='white').place(relx=0.849,rely=0.55,anchor=CENTER)
    
    point.mainloop()








# instruction page:

def instructionpage():
    global instruc
    instruc=tk.Toplevel()
    
    instruc.iconbitmap("icon ICO.ico")
    instruc.title('Monty Hall - Instruction')
    instruc.state('zoomed')
    instruc.protocol('WM_DELETE_WINDOW',instrucclose)

    instrucpic=ImageTk.PhotoImage(Image.open("Front image.png"))
    instrucpanel=Label(instruc,image=instrucpic)
    instrucpanel.pack(side='top',fill='both',expand='yes')


    Label(instruc,text=('''
The Monty Hall problem is a famous probability puzzle named after the host of the television game show "Let's Make a Deal," Monty Hall.

Setup:
Imagine you are a contestant on a game show.
There are three doors (Door A, Door B, and Door C).
Behind one of the doors is a car (the prize you want), and behind the other two doors are goats (something you don't want).

Gameplay:
You initially choose one of the three doors, say Door A.
The host (Monty Hall) knows what is behind each door.
Monty, who wants to make the game interesting, opens one of the other two doors, revealing a goat.
Now, there are two doors left: the one you initially chose (Door A) and the one Monty didn't open.

Decision Point:
Monty gives you a choice: stick with your original choice (Door A) or switch to the other unopened door (let's call it Door B).

Decision Strategies:
"Stick" Strategy: If you decide to stick with your original choice, you will open the door you initially picked and claim whatever is behind it.
"Switch" Strategy: If you decide to switch to the other unopened door, you will open that door and claim whatever is behind it.

Theoretical Solution:
The probability of winning the car is higher if you switch doors rather than sticking with your initial choice.

Explanation:
Initially, there was a 1/3 chance that you picked the car and a 2/3 chance that the car is behind one of the other doors.
When Monty reveals a goat behind one of the other doors, the 2/3 chance is now concentrated behind the unopened door you did not choose.      
                        '''),font=('Arial',15), fg="white",bg='#3c1c0f',justify='left').place(relx=0.5,rely=0.559,anchor=CENTER)

    
    
    instruc.mainloop()







# Home Page:

home=tk.Tk()
home.iconbitmap("icon ICO.ico")
home.title('Monty Hall')
home.configure(bg="black")
home.state('zoomed')
home.protocol('WM_DELETE_WINDOW',homeclose)

homepic=ImageTk.PhotoImage(Image.open("Front image.png"))
homepanel=Label(home,image=homepic)
homepanel.pack(side='top',fill='both',expand='yes')

Button(home,text='Play',font=('Comic Sans MS',14,'bold'),command=playGame,height=1,width=9,bg='#3c1c0f',
       fg='white',activebackground='black',activeforeground='white').place(relx=0.155,rely=0.5,anchor=CENTER)
Button(home,text='Points',font=('Comic Sans MS',14,'bold'),command=pointspage,height=1,width=9,bg='#3c1c0f',
       fg='white',activebackground='black',activeforeground='white').place(relx=0.383,rely=0.5,anchor=CENTER)
Button(home,text='Direction',font=('Comic Sans MS',14,'bold'),command=instructionpage,height=1,width=9,bg='#3c1c0f',
       fg='white',activebackground='black',activeforeground='white').place(relx=0.619,rely=0.5,anchor=CENTER)
Button(home,text='Quit',font=('Comic Sans MS',14,'bold'),command=homelogout,height=1,width=9,bg='#3c1c0f',
       fg='white',activebackground='black',activeforeground='white').place(relx=0.849,rely=0.5,anchor=CENTER)

home.mainloop()



